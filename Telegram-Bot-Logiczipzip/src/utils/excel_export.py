import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


CATEGORY_COLORS = {
    "МТС Физ": "FF6600",
    "МТС Е.": "CC0000",
    "МЕГАФОН": "006633",
    "ТЕЛЕ2": "1A1A2E",
    "БИЛАЙН": "FFD700",
    "YOTA": "0066CC",
    "М/СТ/Т": "993399",
}

DEFAULT_CAT_COLOR = "4472C4"


def _get_cat_header_fill(cat_name: str) -> PatternFill:
    color = CATEGORY_COLORS.get(cat_name, DEFAULT_CAT_COLOR)
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _fmt_price(val) -> str:
    val = float(val or 0)
    if val == int(val):
        return f"{int(val)}$"
    return f"{val}$"


def generate_availability_excel(rows: list[dict], title: str = "Наличие") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    categories = sorted(set(r["category_name"] for r in rows))

    headers = ["Логин:", "Пароль:"]
    header_fills = [header_fill, header_fill]
    for cat in categories:
        headers.append(cat)
        header_fills.append(_get_cat_header_fill(cat))
    headers.append("Сумма:")
    header_fills.append(header_fill)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fills[col_idx - 1]
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    accounts = {}
    account_passwords = {}
    for r in rows:
        phone = r["phone"]
        if phone not in accounts:
            accounts[phone] = {}
            account_passwords[phone] = r.get("password", "")
        remaining = r["effective_max"] - r["used_signatures"]
        real_rev = float(r.get("real_revenue", 0) or 0)
        accounts[phone][r["category_name"]] = {
            "remaining": remaining,
            "max": r["effective_max"],
            "revenue": real_rev,
        }

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    row_idx = 2
    grand_total = 0.0
    for phone in sorted(accounts.keys()):
        cell_phone = ws.cell(row=row_idx, column=1, value=phone)
        cell_phone.border = thin_border
        cell_phone.alignment = Alignment(horizontal="left")

        cell_pass = ws.cell(row=row_idx, column=2, value=account_passwords.get(phone, ""))
        cell_pass.border = thin_border
        cell_pass.alignment = Alignment(horizontal="left")

        row_total = 0.0
        for cat_idx, cat in enumerate(categories):
            col = 3 + cat_idx
            data = accounts[phone].get(cat, {"remaining": 0, "max": 0, "revenue": 0})
            cat_revenue = data["revenue"]
            row_total += cat_revenue
            cell_val = f"{data['remaining']}/{data['max']} - {_fmt_price(cat_revenue)}"
            cell = ws.cell(row=row_idx, column=col, value=cell_val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            cell.fill = green_fill if data["remaining"] > 0 else red_fill

        sum_col = 3 + len(categories)
        cell_sum = ws.cell(row=row_idx, column=sum_col, value=f"{_fmt_price(row_total)}")
        cell_sum.border = thin_border
        cell_sum.alignment = Alignment(horizontal="center")
        cell_sum.font = Font(bold=True)
        grand_total += row_total
        row_idx += 1

    row_idx += 1
    sum_col = 3 + len(categories)
    cell_total = ws.cell(row=row_idx, column=sum_col, value=f"{_fmt_price(grand_total)}")
    cell_total.font = Font(bold=True, size=12)
    cell_total.alignment = Alignment(horizontal="center")
    cell_total.border = thin_border

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="availability_")
    os.close(fd)
    wb.save(path)
    return path


def generate_sales_excel(rows: list[dict], title: str = "Продажи") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    categories = sorted(set(r["category_name"] for r in rows))

    headers = ["Логин:", "Пароль:"]
    header_fills = [header_fill, header_fill]
    for cat in categories:
        headers.append(cat)
        header_fills.append(_get_cat_header_fill(cat))
    headers.append("Сумма:")
    header_fills.append(header_fill)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fills[col_idx - 1]
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    accounts = {}
    account_passwords = {}
    for r in rows:
        phone = r["phone"]
        if phone not in accounts:
            accounts[phone] = {"cats": {}, "total_revenue": 0}
            account_passwords[phone] = r.get("password", "")
        remaining = r["effective_max"] - r["used_signatures"]
        price = r.get("category_price", 0) or 0
        sold = r.get("sold_count", 0) or 0
        revenue = round(r.get("revenue", 0) or 0, 2)
        accounts[phone]["cats"][r["category_name"]] = {
            "remaining": remaining,
            "max": r["effective_max"],
            "price": price,
            "sold": sold,
            "revenue": revenue,
        }
        accounts[phone]["total_revenue"] += revenue

    row_idx = 2
    grand_total = 0.0
    for phone in sorted(accounts.keys()):
        cell_phone = ws.cell(row=row_idx, column=1, value=phone)
        cell_phone.border = thin_border
        cell_phone.alignment = Alignment(horizontal="left")

        cell_pass = ws.cell(row=row_idx, column=2, value=account_passwords.get(phone, ""))
        cell_pass.border = thin_border
        cell_pass.alignment = Alignment(horizontal="left")

        for cat_idx, cat in enumerate(categories):
            col = 3 + cat_idx
            data = accounts[phone]["cats"].get(cat, {"remaining": 0, "max": 0, "price": 0, "sold": 0, "revenue": 0})
            cell_val = f"{data['remaining']}/{data['max']} - {_fmt_price(data['revenue'])}"
            cell = ws.cell(row=row_idx, column=col, value=cell_val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if data.get("sold", 0) > 0:
                cell.fill = yellow_fill
            elif data["remaining"] > 0:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

        sum_col = 3 + len(categories)
        row_revenue = round(accounts[phone]["total_revenue"], 2)
        cell_sum = ws.cell(row=row_idx, column=sum_col, value=f"{_fmt_price(row_revenue)}")
        cell_sum.border = thin_border
        cell_sum.alignment = Alignment(horizontal="center")
        cell_sum.font = Font(bold=True)

        grand_total += row_revenue
        row_idx += 1

    row_idx += 1
    sum_col = 3 + len(categories)
    cell_total = ws.cell(row=row_idx, column=sum_col, value=f"{_fmt_price(round(grand_total, 2))}")
    cell_total.font = Font(bold=True, size=12)
    cell_total.alignment = Alignment(horizontal="center")
    cell_total.border = thin_border

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="sales_")
    os.close(fd)
    wb.save(path)
    return path
